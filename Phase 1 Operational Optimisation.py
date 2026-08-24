"""
Phase 1 Operational Optimisation Model
======================================

This script implements the Phase 1 MILP described in Section 3.5.1 of the
dissertation methodology.

Core assumptions
----------------
- Fixed capacities: PV 160 MW, wind 130 MW, BESS 100 MW / 400 MWh (4 h).
- Hourly 2024 Great Britain data.
- Seven asset configurations.
- Renewable-containing configurations: grid connection = 60-100% of installed RES.
- Standalone BESS: fixed grid connection = BESS power rating.
- Grid import prohibited except for standalone BESS.
- Merchant and fixed-price PPA scenarios are solved separately.
- Standalone BESS is Merchant-only.
- Linear degradation cost is charged on discharged energy.
- No curtailment penalty is included in the objective.
- BESS SOC is bounded between 10% and 90%, starts at 50%, and returns to
  50% at the end of the full Phase 1 horizon.
"""

```
import re
from pathlib import Path

import pandas as pd
import pyomo.environ as pyo
import matplotlib.pyplot as plt
import numpy as np


# ============================================================
# 1. File paths
# ============================================================

PRICE_PATH = Path("data/raw/GB_price.csv")

PV_PATH = Path("data/raw/ninja-pv-GB.csv")

WIND_PATH = Path("data/raw/ninja-wind-GB.csv")

OUTPUT_DIRECTORY = Path("results/phase1")


# ============================================================
# 2. User settings
# ============================================================

START_DATE = "2024-01-01 00:00:00+00:00"
END_DATE = "2025-01-01 00:00:00+00:00"

# If the parsed wind file contains one national column such as GB,
# United Kingdom, or electricity, the script selects it automatically.
#
# If it contains many regional columns, set the exact desired column name
# here after checking the printed list.
WIND_COLUMN = None

# If WIND_COLUMN remains None and multiple regional columns are found:
#   "mean"  -> use the unweighted average as a temporary proxy
#   "error" -> stop and ask you to specify WIND_COLUMN
MULTI_ZONE_WIND_METHOD = "mean"

# ============================================================
# 3. Model parameters
# ============================================================

PV_CAPACITY_MW = 160.0
WIND_CAPACITY_MW = 130.0

BATTERY_ENERGY_MWH = 400.0
BATTERY_POWER_MW = 100.0

SOC_INITIAL = 0.50
SOC_MIN = 0.10
SOC_MAX = 0.90

ETA_CHARGE = 0.95
ETA_DISCHARGE = 0.95

PPA_PRICE_GBP_PER_MWH = 86.21
DEGRADATION_COST_GBP_PER_MWH = 42.80

DELTA_T = 1.0

# ============================================================
# 4. Data import and preprocessing
# ============================================================

print("\nLoading input data...")

# ------------------------------------------------------------
# 4.1 Day-ahead electricity prices
# ------------------------------------------------------------

price_data = pd.read_csv(PRICE_PATH)

# Parse timestamps as UTC
price_data["time"] = pd.to_datetime(
    price_data["start_time"],
    utc=True,
    errors="coerce"
)

price_data["price"] = pd.to_numeric(
    price_data["price"],
    errors="coerce"
)

# Keep only required columns
price_data = price_data[
    ["time", "price"]
].copy()

# Remove invalid observations
price_data = price_data.dropna(
    subset=["time", "price"]
)

# Remove duplicate timestamps if any
price_data = price_data.drop_duplicates(
    subset="time",
    keep="first"
)

# Sort chronologically because the original price file
# is ordered from newest to oldest
price_data = price_data.sort_values(
    "time"
)

price_data = price_data.set_index("time")


# ------------------------------------------------------------
# 4.2 PV capacity factor
# ------------------------------------------------------------
#
# Renewables.ninja files contain three metadata rows before
# the actual column names:
#
# row 1 = dataset description
# row 2 = units
# row 3 = metadata dictionary
# row 4 = actual header ("time", "NATIONAL", ...)
#
# Therefore skiprows=3.
# ------------------------------------------------------------

pv_data = pd.read_csv(
    PV_PATH,
    skiprows=3
)

# Clean column names
pv_data.columns = [
    str(col).strip()
    for col in pv_data.columns
]

print("\nPV columns:")
print(pv_data.columns.tolist())

if "time" not in pv_data.columns:
    raise ValueError(
        "PV file does not contain a 'time' column "
        "after skipping the metadata rows."
    )

if "NATIONAL" not in pv_data.columns:
    raise ValueError(
        "PV file does not contain a 'NATIONAL' column."
    )

pv_data["time"] = pd.to_datetime(
    pv_data["time"],
    utc=True,
    errors="coerce"
)

pv_data["pv_cf"] = pd.to_numeric(
    pv_data["NATIONAL"],
    errors="coerce"
)

pv_data = pv_data[
    ["time", "pv_cf"]
].copy()

pv_data = pv_data.dropna(
    subset=["time", "pv_cf"]
)

pv_data = pv_data.drop_duplicates(
    subset="time",
    keep="first"
)

pv_data = pv_data.sort_values(
    "time"
)

pv_data = pv_data.set_index("time")


# ------------------------------------------------------------
# 4.3 Wind capacity factor
# ------------------------------------------------------------

wind_data = pd.read_csv(
    WIND_PATH,
    skiprows=3
)

wind_data.columns = [
    str(col).strip()
    for col in wind_data.columns
]

print("\nWind columns:")
print(wind_data.columns.tolist())

if "time" not in wind_data.columns:
    raise ValueError(
        "Wind file does not contain a 'time' column "
        "after skipping the metadata rows."
    )

wind_data["time"] = pd.to_datetime(
    wind_data["time"],
    utc=True,
    errors="coerce"
)

# ------------------------------------------------------------
# Select wind column
#
# Priority:
# 1. User explicitly specifies WIND_COLUMN
# 2. NATIONAL
# 3. GB
# 4. United Kingdom
# 5. electricity
# 6. Regional mean, only if requested
# ------------------------------------------------------------

if WIND_COLUMN is not None:

    if WIND_COLUMN not in wind_data.columns:
        raise ValueError(
            f"WIND_COLUMN='{WIND_COLUMN}' was not found.\n"
            f"Available wind columns:\n"
            f"{wind_data.columns.tolist()}"
        )

    selected_wind_column = WIND_COLUMN

else:

    national_candidates = [
        "NATIONAL",
        "GB",
        "United Kingdom",
        "electricity",
    ]

    selected_wind_column = None

    for candidate in national_candidates:

        if candidate in wind_data.columns:
            selected_wind_column = candidate
            break

    # If no national column can be found:
    if selected_wind_column is None:

        regional_columns = [
            col
            for col in wind_data.columns
            if col != "time"
        ]

        if MULTI_ZONE_WIND_METHOD == "mean":

            print(
                "\nWarning: No national wind column found. "
                "Using the unweighted mean of regional columns."
            )

            for col in regional_columns:
                wind_data[col] = pd.to_numeric(
                    wind_data[col],
                    errors="coerce"
                )

            wind_data["wind_cf"] = (
                wind_data[
                    regional_columns
                ].mean(axis=1)
            )

        else:

            raise ValueError(
                "Multiple regional wind columns were found "
                "but no NATIONAL column could be identified. "
                "Please specify WIND_COLUMN."
            )


# If a national/specified column was found
if selected_wind_column is not None:

    print(
        f"\nSelected wind column: "
        f"{selected_wind_column}"
    )

    wind_data["wind_cf"] = pd.to_numeric(
        wind_data[selected_wind_column],
        errors="coerce"
    )


wind_data = wind_data[
    ["time", "wind_cf"]
].copy()

wind_data = wind_data.dropna(
    subset=["time", "wind_cf"]
)

wind_data = wind_data.drop_duplicates(
    subset="time",
    keep="first"
)

wind_data = wind_data.sort_values(
    "time"
)

wind_data = wind_data.set_index("time")


# ------------------------------------------------------------
# 4.4 Restrict all datasets to modelling period
# ------------------------------------------------------------

start_time = pd.Timestamp(START_DATE)
end_time = pd.Timestamp(END_DATE)

price_data = price_data.loc[
    (price_data.index >= start_time)
    & (price_data.index < end_time)
].copy()

pv_data = pv_data.loc[
    (pv_data.index >= start_time)
    & (pv_data.index < end_time)
].copy()

wind_data = wind_data.loc[
    (wind_data.index >= start_time)
    & (wind_data.index < end_time)
].copy()


# ------------------------------------------------------------
# 4.5 Merge price, PV and wind data
# ------------------------------------------------------------

model_data = (
    price_data
    .join(
        pv_data,
        how="inner"
    )
    .join(
        wind_data,
        how="inner"
    )
)

model_data = model_data[
    ["price", "pv_cf", "wind_cf"]
].copy()

model_data = model_data.dropna()

model_data = model_data.sort_index()


# ------------------------------------------------------------
# 4.6 Data validation
# ------------------------------------------------------------

if model_data.empty:
    raise ValueError(
        "model_data is empty after merging. "
        "Check the dates and timestamps."
    )

# Capacity factors should normally lie between 0 and 1.
# Slight numerical deviations are clipped.
model_data["pv_cf"] = model_data[
    "pv_cf"
].clip(
    lower=0.0,
    upper=1.0
)

model_data["wind_cf"] = model_data[
    "wind_cf"
].clip(
    lower=0.0,
    upper=1.0
)


# ------------------------------------------------------------
# 4.7 Check temporal continuity
# ------------------------------------------------------------

expected_index = pd.date_range(
    start=start_time,
    end=end_time,
    freq="h",
    inclusive="left"
)

missing_hours = expected_index.difference(
    model_data.index
)

print("\n================================================")
print("MODEL DATA SUMMARY")
print("================================================")

print(
    f"Start: {model_data.index.min()}"
)

print(
    f"End:   {model_data.index.max()}"
)

print(
    f"Number of observations: "
    f"{len(model_data)}"
)

print(
    f"Expected observations: "
    f"{len(expected_index)}"
)

print(
    f"Missing hourly observations: "
    f"{len(missing_hours)}"
)

if len(missing_hours) > 0:

    print("\nFirst missing timestamps:")

    print(
        missing_hours[:10]
    )


print("\nSummary statistics:")
print(
    model_data[
        ["price", "pv_cf", "wind_cf"]
    ].describe()
)


print("\nFirst five observations:")
print(
    model_data.head()
)


print("\nLast five observations:")
print(
    model_data.tail()
)

# ============================================================
# 4. Scenario settings
# ============================================================

# Contract indicators for separate Merchant and PPA scenarios
MERCHANT_ALPHA = 0.0
PPA_ALPHA = 1.0

GRID_LEVELS = [0.60, 0.70, 0.80, 0.90, 1.00]

CONFIGURATIONS = {
    "PV": {
        "pv": 1,
        "wind": 0,
        "bess": 0,
        "grid_import": 0,
    },
    "Wind": {
        "pv": 0,
        "wind": 1,
        "bess": 0,
        "grid_import": 0,
    },
    "PV_Wind": {
        "pv": 1,
        "wind": 1,
        "bess": 0,
        "grid_import": 0,
    },
    "BESS": {
        "pv": 0,
        "wind": 0,
        "bess": 1,
        "grid_import": 1,   # Only standalone BESS can import
    },
    "PV_BESS": {
        "pv": 1,
        "wind": 0,
        "bess": 1,
        "grid_import": 0,
    },
    "Wind_BESS": {
        "pv": 0,
        "wind": 1,
        "bess": 1,
        "grid_import": 0,
    },
    "PV_Wind_BESS": {
        "pv": 1,
        "wind": 1,
        "bess": 1,
        "grid_import": 0,
    },
}


# ============================================================
# 5. Required input data
# ============================================================

# This optimisation section assumes that your earlier data-processing
# code has already produced a DataFrame called model_data with:
#
#   model_data.index       -> hourly DatetimeIndex
#   model_data["price"]     -> day-ahead electricity price (£/MWh)
#   model_data["pv_cf"]     -> PV capacity factor (0-1)
#   model_data["wind_cf"]   -> wind capacity factor (0-1)
#
# For example:
#
# model_data = pd.DataFrame({
#     "price": price_series,
#     "pv_cf": pv_cf_series,
#     "wind_cf": wind_cf_series,
# }).dropna()
#
# model_data = model_data.loc[START_DATE:END_DATE]

required_columns = {"price", "pv_cf", "wind_cf"}

missing_columns = required_columns.difference(model_data.columns)

if missing_columns:
    raise ValueError(
        f"model_data is missing required columns: {missing_columns}"
    )

model_data = model_data.copy().sort_index()

# Restrict model period
model_data = model_data.loc[
    (model_data.index >= pd.Timestamp(START_DATE))
    & (model_data.index < pd.Timestamp(END_DATE))
].copy()

model_data = model_data.dropna(
    subset=["price", "pv_cf", "wind_cf"]
)

if model_data.empty:
    raise ValueError("No observations remain after date filtering.")

print(f"Model period: {model_data.index.min()} to {model_data.index.max()}")
print(f"Number of hourly observations: {len(model_data)}")


# ============================================================
# 6. Helper functions
# ============================================================

def get_installed_res_capacity(config_name):
    """
    Installed renewable capacity for each configuration.

    Standalone BESS has P_RES = 0 by definition.
    """

    flags = CONFIGURATIONS[config_name]

    p_res = (
        flags["pv"] * PV_CAPACITY_MW
        + flags["wind"] * WIND_CAPACITY_MW
    )

    return float(p_res)


def get_grid_connection_capacity(config_name, grid_level):
    """
    Grid connection capacity for each scenario.

    Renewable configurations:
        P_grid = gamma * P_RES

    Standalone BESS:
        P_RES = 0,
        but its original standalone grid connection definition is retained,
        with P_grid = BATTERY_POWER_MW.
    """

    if config_name == "BESS":
        return float(BATTERY_POWER_MW)

    p_res = get_installed_res_capacity(config_name)

    return float(grid_level * p_res)


def build_generation_profiles(config_name, data):
    """
    Create renewable generation profiles (MW) according to configuration.
    """

    flags = CONFIGURATIONS[config_name]

    pv_generation = (
        flags["pv"]
        * PV_CAPACITY_MW
        * data["pv_cf"].to_numpy()
    )

    wind_generation = (
        flags["wind"]
        * WIND_CAPACITY_MW
        * data["wind_cf"].to_numpy()
    )

    res_generation = pv_generation + wind_generation

    return (
        pv_generation,
        wind_generation,
        res_generation,
    )


def calculate_curtailment_breakdown(
    res_generation,
    grid_export,
    battery_charge,
    battery_discharge,
    grid_capacity,
):
    """
    Post-process renewable curtailment into physical (grid-induced)
    and economic components without changing the optimisation dispatch.

    Notes
    -----
    - grid_export is total export, including battery discharge.
    - renewable direct export is therefore grid_export - battery discharge.
    - physical curtailment is the renewable generation that could not be
      accommodated after battery charging because the remaining shared
      grid headroom was insufficient.
    - economic curtailment is the residual unutilised renewable generation
      after physical curtailment is removed.
    """

    res_generation = np.asarray(res_generation, dtype=float)
    grid_export = np.asarray(grid_export, dtype=float)
    battery_charge = np.asarray(battery_charge, dtype=float)
    battery_discharge = np.asarray(battery_discharge, dtype=float)

    renewable_export = np.maximum(
        0.0,
        grid_export - battery_discharge,
    )

    total_curtailment = np.maximum(
        0.0,
        res_generation
        - battery_charge
        - renewable_export,
    )

    grid_headroom_for_res = np.maximum(
        0.0,
        grid_capacity - battery_discharge,
    )

    physical_curtailment = np.maximum(
        0.0,
        res_generation
        - battery_charge
        - grid_headroom_for_res,
    )

    # Numerical safeguard: physical curtailment is one component
    # of total curtailment and therefore cannot exceed it.
    physical_curtailment = np.minimum(
        physical_curtailment,
        total_curtailment,
    )

    economic_curtailment = np.maximum(
        0.0,
        total_curtailment - physical_curtailment,
    )

    return {
        "renewable_export": renewable_export,
        "grid_headroom_for_res": grid_headroom_for_res,
        "total_curtailment": total_curtailment,
        "physical_curtailment": physical_curtailment,
        "economic_curtailment": economic_curtailment,
    }


# ============================================================
# 7. Unified optimisation model
# ============================================================

def run_scenario(
    config_name,
    grid_level,
    alpha=MERCHANT_ALPHA,
    solver_name="highs",
):

    flags = CONFIGURATIONS[config_name]

    pv_enabled = flags["pv"]
    wind_enabled = flags["wind"]
    bess_enabled = flags["bess"]
    grid_import_enabled = flags["grid_import"]

    # --------------------------------------------------------
    # Scenario-specific capacities
    # --------------------------------------------------------

    p_res_installed = get_installed_res_capacity(config_name)

    p_grid = get_grid_connection_capacity(
        config_name,
        grid_level,
    )

    (
        pv_generation,
        wind_generation,
        res_generation,
    ) = build_generation_profiles(
        config_name,
        model_data,
    )

    prices = model_data["price"].to_numpy()

    n_hours = len(model_data)

    # --------------------------------------------------------
    # Create model
    # --------------------------------------------------------

    m = pyo.ConcreteModel()

    m.T = pyo.RangeSet(0, n_hours - 1)

    # --------------------------------------------------------
    # Parameters
    # --------------------------------------------------------

    m.price = pyo.Param(
        m.T,
        initialize={
            t: float(prices[t])
            for t in range(n_hours)
        }
    )

    m.pv_generation = pyo.Param(
        m.T,
        initialize={
            t: float(pv_generation[t])
            for t in range(n_hours)
        }
    )

    m.wind_generation = pyo.Param(
        m.T,
        initialize={
            t: float(wind_generation[t])
            for t in range(n_hours)
        }
    )

    m.res_generation = pyo.Param(
        m.T,
        initialize={
            t: float(res_generation[t])
            for t in range(n_hours)
        }
    )

    # --------------------------------------------------------
    # Decision variables
    # --------------------------------------------------------

    # Grid export
    m.P_out = pyo.Var(
        m.T,
        domain=pyo.NonNegativeReals
    )

    # Grid import
    m.P_in = pyo.Var(
        m.T,
        domain=pyo.NonNegativeReals
    )

    # Battery charging
    m.C = pyo.Var(
        m.T,
        domain=pyo.NonNegativeReals
    )

    # Battery discharging
    m.D = pyo.Var(
        m.T,
        domain=pyo.NonNegativeReals
    )

    # Renewable curtailment
    m.Curt = pyo.Var(
        m.T,
        domain=pyo.NonNegativeReals
    )

    # Battery energy level
    m.E = pyo.Var(
        m.T,
        domain=pyo.NonNegativeReals
    )

    # Binary variable:
    # 1 = charging
    # 0 = discharging / idle
    m.battery_mode = pyo.Var(
        m.T,
        domain=pyo.Binary
    )

    # --------------------------------------------------------
    # Unified objective function
    # --------------------------------------------------------

    def objective_rule(m):

        return sum(

            (
                alpha
                * PPA_PRICE_GBP_PER_MWH
                * m.P_out[t]

                +

                (1 - alpha)
                * m.price[t]
                * m.P_out[t]

                -

                m.price[t]
                * m.P_in[t]

                -

                DEGRADATION_COST_GBP_PER_MWH
                * m.D[t]
            )

            * DELTA_T

            for t in m.T
        )

    m.objective = pyo.Objective(
        rule=objective_rule,
        sense=pyo.maximize
    )

    # --------------------------------------------------------
    # 7.1 Grid export limit
    # --------------------------------------------------------

    def export_limit_rule(m, t):

        return m.P_out[t] <= p_grid

    m.export_limit = pyo.Constraint(
        m.T,
        rule=export_limit_rule
    )

    # --------------------------------------------------------
    # 7.2 Grid import rule
    #
    # Only standalone BESS may import electricity.
    # All other configurations:
    # P_in = 0
    # --------------------------------------------------------

    def import_limit_rule(m, t):

        if grid_import_enabled == 1:
            return m.P_in[t] <= p_grid

        return m.P_in[t] == 0

    m.import_limit = pyo.Constraint(
        m.T,
        rule=import_limit_rule
    )

    # --------------------------------------------------------
    # 7.3 Renewable curtailment limit
    #
    # Curtailment cannot exceed renewable generation.
    # For BESS-alone RES generation = 0,
    # therefore Curt = 0 automatically.
    # --------------------------------------------------------

    def curtailment_limit_rule(m, t):

        return m.Curt[t] <= m.res_generation[t]

    m.curtailment_limit = pyo.Constraint(
        m.T,
        rule=curtailment_limit_rule
    )

    # --------------------------------------------------------
    # 7.4 Battery charging/discharging limits
    # --------------------------------------------------------

    def charge_limit_rule(m, t):

        if bess_enabled == 0:
            return m.C[t] == 0

        return (
            m.C[t]
            <= BATTERY_POWER_MW
            * m.battery_mode[t]
        )

    m.charge_limit = pyo.Constraint(
        m.T,
        rule=charge_limit_rule
    )

    def discharge_limit_rule(m, t):

        if bess_enabled == 0:
            return m.D[t] == 0

        return (
            m.D[t]
            <= BATTERY_POWER_MW
            * (1 - m.battery_mode[t])
        )

    m.discharge_limit = pyo.Constraint(
        m.T,
        rule=discharge_limit_rule
    )

    # --------------------------------------------------------
    # 7.5 Standalone BESS import definition
    #
    # For BESS-alone:
    # grid imports are used only for battery charging:
    #
    # P_in = C
    #
    # For all other scenarios:
    # P_in = 0 already enforced above.
    # --------------------------------------------------------

    def bess_import_definition_rule(m, t):

        if config_name == "BESS":
            return m.P_in[t] == m.C[t]

        return pyo.Constraint.Skip

    m.bess_import_definition = pyo.Constraint(
        m.T,
        rule=bess_import_definition_rule
    )

    # --------------------------------------------------------
    # 7.6 Energy balance
    #
    # RES + grid import + battery discharge
    # =
    # grid export + battery charge + curtailment
    # --------------------------------------------------------

    def energy_balance_rule(m, t):

        return (
            m.res_generation[t]
            + m.P_in[t]
            + m.D[t]
            ==
            m.P_out[t]
            + m.C[t]
            + m.Curt[t]
        )

    m.energy_balance = pyo.Constraint(
        m.T,
        rule=energy_balance_rule
    )

    # --------------------------------------------------------
    # 7.7 Battery SOC bounds
    # --------------------------------------------------------

    E_MIN = SOC_MIN * BATTERY_ENERGY_MWH
    E_MAX = SOC_MAX * BATTERY_ENERGY_MWH
    E_INITIAL = SOC_INITIAL * BATTERY_ENERGY_MWH

    def energy_min_rule(m, t):

        if bess_enabled == 0:
            return m.E[t] == 0

        return m.E[t] >= E_MIN

    m.energy_min = pyo.Constraint(
        m.T,
        rule=energy_min_rule
    )

    def energy_max_rule(m, t):

        if bess_enabled == 0:
            return pyo.Constraint.Skip

        return m.E[t] <= E_MAX

    m.energy_max = pyo.Constraint(
        m.T,
        rule=energy_max_rule
    )

    # --------------------------------------------------------
    # 7.8 Battery SOC evolution
    # --------------------------------------------------------

    def soc_balance_rule(m, t):

        if bess_enabled == 0:
            return pyo.Constraint.Skip

        if t == 0:

            return (
                m.E[t]
                ==
                E_INITIAL
                + ETA_CHARGE * m.C[t] * DELTA_T
                - (
                    m.D[t]
                    / ETA_DISCHARGE
                )
                * DELTA_T
            )

        return (
            m.E[t]
            ==
            m.E[t - 1]
            + ETA_CHARGE * m.C[t] * DELTA_T
            - (
                m.D[t]
                / ETA_DISCHARGE
            )
            * DELTA_T
        )

    m.soc_balance = pyo.Constraint(
        m.T,
        rule=soc_balance_rule
    )

    # --------------------------------------------------------
    # 7.9 Terminal SOC condition
    #
    # Battery SOC is continuous across the full modelling
    # horizon, but the final energy level is required to
    # return to the initial SOC to avoid end-of-horizon bias.
    # --------------------------------------------------------

    def terminal_soc_rule(m):

        if bess_enabled == 0:
            return pyo.Constraint.Skip

        final_t = n_hours - 1

        return m.E[final_t] == E_INITIAL

    m.terminal_soc = pyo.Constraint(
        rule=terminal_soc_rule
    )

    # --------------------------------------------------------
    # Solve
    # --------------------------------------------------------

    solver = pyo.SolverFactory(solver_name)

    results = solver.solve(
        m,
        tee=False
    )

    termination = (
        results.solver.termination_condition
    )

    if termination != pyo.TerminationCondition.optimal:

        print(
            f"Warning: {config_name}, "
            f"{grid_level:.0%}: "
            f"{termination}"
        )

    # --------------------------------------------------------
    # Extract hourly results
    # --------------------------------------------------------

    hourly = pd.DataFrame(
        index=model_data.index
    )

    hourly["price"] = prices

    hourly["PV_generation_MW"] = pv_generation
    hourly["Wind_generation_MW"] = wind_generation
    hourly["RES_generation_MW"] = res_generation

    hourly["P_out_MW"] = [
        pyo.value(m.P_out[t])
        for t in m.T
    ]

    hourly["P_in_MW"] = [
        pyo.value(m.P_in[t])
        for t in m.T
    ]

    hourly["Charge_MW"] = [
        pyo.value(m.C[t])
        for t in m.T
    ]

    hourly["Discharge_MW"] = [
        pyo.value(m.D[t])
        for t in m.T
    ]

    hourly["Curtailment_MW"] = [
        pyo.value(m.Curt[t])
        for t in m.T
    ]

    hourly["Battery_energy_MWh"] = [
        pyo.value(m.E[t])
        for t in m.T
    ]

    if bess_enabled == 1:

        hourly["SOC"] = (
            hourly["Battery_energy_MWh"]
            / BATTERY_ENERGY_MWH
        )

    else:

        hourly["SOC"] = 0.0

    # --------------------------------------------------------
    # Curtailment decomposition (post-processing)
    #
    # The optimisation variable m.Curt is retained as total
    # unutilised renewable generation. Here it is decomposed
    # into:
    #   1. physical / grid-induced curtailment, and
    #   2. economic curtailment.
    #
    # This does NOT change dispatch or revenue.
    # --------------------------------------------------------

    curtailment_breakdown = calculate_curtailment_breakdown(
        res_generation=hourly["RES_generation_MW"].to_numpy(),
        grid_export=hourly["P_out_MW"].to_numpy(),
        battery_charge=hourly["Charge_MW"].to_numpy(),
        battery_discharge=hourly["Discharge_MW"].to_numpy(),
        grid_capacity=p_grid,
    )

    hourly["Renewable_export_MW"] = (
        curtailment_breakdown["renewable_export"]
    )

    hourly["Grid_headroom_for_RES_MW"] = (
        curtailment_breakdown["grid_headroom_for_res"]
    )

    hourly["Total_curtailment_MW"] = (
        curtailment_breakdown["total_curtailment"]
    )

    hourly["Physical_curtailment_MW"] = (
        curtailment_breakdown["physical_curtailment"]
    )

    hourly["Economic_curtailment_MW"] = (
        curtailment_breakdown["economic_curtailment"]
    )

    # Check that the post-processed total agrees with the
    # optimisation model's original curtailment variable.
    if not np.allclose(
        hourly["Total_curtailment_MW"].to_numpy(),
        hourly["Curtailment_MW"].to_numpy(),
        atol=1e-5,
        rtol=1e-7,
    ):
        raise ValueError(
            "Curtailment post-processing does not match "
            "the optimisation energy balance."
        )

    # --------------------------------------------------------
    # Revenue components
    # --------------------------------------------------------

    hourly["Merchant_export_revenue_GBP"] = (
        hourly["price"]
        * hourly["P_out_MW"]
        * DELTA_T
    )

    hourly["Grid_import_cost_GBP"] = (
        hourly["price"]
        * hourly["P_in_MW"]
        * DELTA_T
    )

    hourly["Degradation_cost_GBP"] = (
        DEGRADATION_COST_GBP_PER_MWH
        * hourly["Discharge_MW"]
        * DELTA_T
    )

    hourly["Net_revenue_GBP"] = (

        (1 - alpha)
        * hourly["Merchant_export_revenue_GBP"]

        +

        alpha
        * PPA_PRICE_GBP_PER_MWH
        * hourly["P_out_MW"]
        * DELTA_T

        -

        hourly["Grid_import_cost_GBP"]

        -

        hourly["Degradation_cost_GBP"]
    )

    # --------------------------------------------------------
    # Scenario summary
    # --------------------------------------------------------

    total_res_generation = (
        hourly["RES_generation_MW"].sum()
        * DELTA_T
    )

    total_export = (
        hourly["P_out_MW"].sum()
        * DELTA_T
    )

    total_import = (
        hourly["P_in_MW"].sum()
        * DELTA_T
    )

    total_charge = (
        hourly["Charge_MW"].sum()
        * DELTA_T
    )

    total_discharge = (
        hourly["Discharge_MW"].sum()
        * DELTA_T
    )

    total_curtailment = (
        hourly["Total_curtailment_MW"].sum()
        * DELTA_T
    )

    physical_curtailment = (
        hourly["Physical_curtailment_MW"].sum()
        * DELTA_T
    )

    economic_curtailment = (
        hourly["Economic_curtailment_MW"].sum()
        * DELTA_T
    )

    total_revenue = hourly[
        "Net_revenue_GBP"
    ].sum()

    # Curtailment decomposition must balance annually.
    if not np.isclose(
        total_curtailment,
        physical_curtailment + economic_curtailment,
        atol=1e-5,
        rtol=1e-7,
    ):
        raise ValueError(
            "Annual curtailment decomposition does not balance."
        )

    if total_res_generation > 0:

        total_curtailment_rate = (
            total_curtailment
            / total_res_generation
        )

        physical_curtailment_rate = (
            physical_curtailment
            / total_res_generation
        )

        economic_curtailment_rate = (
            economic_curtailment
            / total_res_generation
        )

    else:

        total_curtailment_rate = np.nan
        physical_curtailment_rate = np.nan
        economic_curtailment_rate = np.nan

    summary = {
        "contract": (
    "PPA"
    if alpha == 1.0
    else "Merchant"
),
        "alpha": alpha,
        "configuration": config_name,

        "grid_level": grid_level,

        "PV_capacity_MW": (
            PV_CAPACITY_MW
            if pv_enabled
            else 0.0
        ),

        "Wind_capacity_MW": (
            WIND_CAPACITY_MW
            if wind_enabled
            else 0.0
        ),

        "RES_capacity_MW": p_res_installed,

        "BESS_power_MW": (
            BATTERY_POWER_MW
            if bess_enabled
            else 0.0
        ),

        "BESS_energy_MWh": (
            BATTERY_ENERGY_MWH
            if bess_enabled
            else 0.0
        ),

        "Grid_connection_MW": p_grid,

        "Grid_import_allowed": (
            bool(grid_import_enabled)
        ),

        "RES_generation_MWh": (
            total_res_generation
        ),

        "Grid_export_MWh": (
            total_export
        ),

        "Grid_import_MWh": (
            total_import
        ),

        "Battery_charge_MWh": (
            total_charge
        ),

        "Battery_discharge_MWh": (
            total_discharge
        ),

        "Physical_curtailment_MWh": (
            physical_curtailment
        ),

        "Economic_curtailment_MWh": (
            economic_curtailment
        ),

        "Total_curtailment_MWh": (
            total_curtailment
        ),

        "Physical_curtailment_rate": (
            physical_curtailment_rate
        ),

        "Economic_curtailment_rate": (
            economic_curtailment_rate
        ),

        "Total_curtailment_rate": (
            total_curtailment_rate
        ),

        "Net_revenue_GBP": (
            total_revenue
        ),

        "Solver_status": (
            str(termination)
        ),
    }

    if total_res_generation > 0:
        print(
            f"    Curtailment -> "
            f"Physical: {physical_curtailment_rate * 100:.2f}% | "
            f"Economic: {economic_curtailment_rate * 100:.2f}% | "
            f"Total: {total_curtailment_rate * 100:.2f}%"
        )

    return summary, hourly


# ============================================================
# 8. Run Phase 1 Merchant scenarios
# ============================================================
#
# Methodology alignment:
# - Renewable-containing configurations are evaluated at
#   60%, 70%, 80%, 90%, and 100% of installed RES capacity.
# - Standalone BESS has one fixed grid connection equal to
#   its battery power rating, so it is solved only once.
# ============================================================

OUTPUT_DIRECTORY.mkdir(
    parents=True,
    exist_ok=True
)

merchant_summary_rows = []
hourly_results = {}

merchant_scenarios = []

for config_name in CONFIGURATIONS:

    if config_name == "BESS":
        merchant_scenarios.append((config_name, 1.00))
    else:
        merchant_scenarios.extend(
            (config_name, grid_level)
            for grid_level in GRID_LEVELS
        )

for scenario_number, (config_name, grid_level) in enumerate(
    merchant_scenarios,
    start=1,
):

    print(
        f"Running Merchant scenario "
        f"{scenario_number}/{len(merchant_scenarios)}: "
        f"{config_name}, "
        f"grid level = {grid_level:.0%}"
    )

    summary, hourly = run_scenario(
        config_name=config_name,
        grid_level=grid_level,
        alpha=MERCHANT_ALPHA,
        solver_name="highs",
    )

    merchant_summary_rows.append(summary)

    scenario_key = (
        f"Merchant_"
        f"{config_name}_"
        f"{int(grid_level * 100)}pct"
    )

    hourly_results[scenario_key] = hourly

    hourly.to_csv(
        OUTPUT_DIRECTORY / f"{scenario_key}.csv"
    )


# ============================================================
# 9. Save Merchant scenario summary
# ============================================================

merchant_summary = pd.DataFrame(
    merchant_summary_rows
)

merchant_summary = merchant_summary.sort_values(
    by=[
        "configuration",
        "grid_level",
    ]
).reset_index(drop=True)

merchant_summary.to_csv(
    OUTPUT_DIRECTORY
    / "merchant_scenario_summary.csv",
    index=False
)

print("\nMerchant optimisation completed.")
print(
    f"Number of scenarios: "
    f"{len(merchant_summary)}"
)

print("\nSummary:")
print(
    merchant_summary[
        [
            "configuration",
            "grid_level",
            "RES_capacity_MW",
            "Grid_connection_MW",
            "Total_curtailment_rate",
            "Grid_export_MWh",
            "Battery_discharge_MWh",
            "Net_revenue_GBP",
        ]
    ]
)


# ============================================================
# 10. Optional checks
# ============================================================

# Check that only standalone BESS imports electricity
import_check = (
    merchant_summary
    .groupby("configuration")[
        "Grid_import_MWh"
    ]
    .max()
)

print("\nMaximum grid import by configuration:")
print(import_check)

# Check grid connection levels
print("\nGrid connection capacities:")
print(
    merchant_summary[
        [
            "configuration",
            "grid_level",
            "RES_capacity_MW",
            "Grid_connection_MW",
        ]
    ]
)

# ============================================================
# 11. Run Phase 1 fixed-price PPA scenarios
# ============================================================
#
# Standalone BESS is excluded because the methodology models
# it exclusively under the Merchant case.
# ============================================================

ppa_summary_rows = []
ppa_hourly_results = {}

ppa_scenarios = [
    (config_name, grid_level)
    for config_name in CONFIGURATIONS
    if config_name != "BESS"
    for grid_level in GRID_LEVELS
]

for scenario_number, (config_name, grid_level) in enumerate(
    ppa_scenarios,
    start=1,
):

    print(
        f"Running PPA scenario "
        f"{scenario_number}/{len(ppa_scenarios)}: "
        f"{config_name}, "
        f"grid level = {grid_level:.0%}"
    )

    summary, hourly = run_scenario(
        config_name=config_name,
        grid_level=grid_level,
        alpha=PPA_ALPHA,
        solver_name="highs",
    )

    ppa_summary_rows.append(summary)

    scenario_key = (
        f"PPA_"
        f"{config_name}_"
        f"{int(grid_level * 100)}pct"
    )

    ppa_hourly_results[scenario_key] = hourly

    hourly.to_csv(
        OUTPUT_DIRECTORY / f"{scenario_key}.csv"
    )


# ============================================================
# 12. Save PPA scenario summary
# ============================================================

ppa_summary = pd.DataFrame(
    ppa_summary_rows
)

ppa_summary = ppa_summary.sort_values(
    by=[
        "configuration",
        "grid_level",
    ]
).reset_index(drop=True)


# ------------------------------------------------------------
# Recalculate summary net revenue explicitly for PPA
#
# This makes the saved summary fully transparent and avoids
# relying only on the value generated inside run_scenario().
# ------------------------------------------------------------

for i, row in ppa_summary.iterrows():

    config_name = row["configuration"]
    grid_level = row["grid_level"]

    scenario_key = (
        f"PPA_"
        f"{config_name}_"
        f"{int(grid_level * 100)}pct"
    )

    hourly = ppa_hourly_results[
        scenario_key
    ]

    ppa_summary.loc[
        i,
        "Net_revenue_GBP"
    ] = hourly[
        "Net_revenue_GBP"
    ].sum()


ppa_summary.to_csv(
    OUTPUT_DIRECTORY
    / "ppa_scenario_summary.csv",
    index=False
)

print("\nPPA optimisation completed.")

print(
    f"Number of PPA scenarios: "
    f"{len(ppa_summary)}"
)

print("\nPPA Summary:")

print(
    ppa_summary[
        [
            "configuration",
            "grid_level",
            "RES_capacity_MW",
            "Grid_connection_MW",
            "Total_curtailment_rate",
            "Grid_export_MWh",
            "Battery_discharge_MWh",
            "Net_revenue_GBP",
        ]
    ]
)


# ============================================================
# 13. PPA validation checks
# ============================================================

# Check that only standalone BESS imports electricity
ppa_import_check = (
    ppa_summary
    .groupby("configuration")[
        "Grid_import_MWh"
    ]
    .max()
)

print(
    "\nMaximum grid import by configuration "
    "under PPA:"
)

print(
    ppa_import_check
)


# ============================================================
# 14. Combine Merchant and PPA summaries
# ============================================================

combined_summary = pd.concat(
    [
        merchant_summary,
        ppa_summary,
    ],
    ignore_index=True
)

combined_summary = combined_summary.sort_values(
    by=[
        "configuration",
        "grid_level",
        "contract",
    ]
).reset_index(drop=True)

combined_summary.to_csv(
    OUTPUT_DIRECTORY
    / "combined_scenario_summary.csv",
    index=False
)

print("\n==============================================")
print("ALL PHASE 1 SCENARIOS COMPLETED")
print("==============================================")

print(
    f"Merchant scenarios: "
    f"{len(merchant_summary)}"
)

print(
    f"PPA scenarios: "
    f"{len(ppa_summary)}"
)

print(
    f"Total scenario rows: "
    f"{len(combined_summary)}"
)


# ============================================================
# 15. Merchant vs PPA comparison
# ============================================================

comparison = combined_summary.pivot_table(
    index=[
        "configuration",
        "grid_level",
        "Grid_connection_MW",
    ],
    columns="contract",
    values="Net_revenue_GBP",
    aggfunc="first",
).reset_index()

if (
    "Merchant" in comparison.columns
    and "PPA" in comparison.columns
):

    comparison[
        "PPA_minus_Merchant_GBP"
    ] = (
        comparison["PPA"]
        - comparison["Merchant"]
    )

    comparison[
        "PPA_vs_Merchant_pct"
    ] = np.where(
        comparison["Merchant"] != 0,

        (
            comparison["PPA"]
            - comparison["Merchant"]
        )
        / comparison["Merchant"]
        * 100,

        np.nan
    )


comparison.to_csv(
    OUTPUT_DIRECTORY
    / "merchant_vs_ppa_comparison.csv",
    index=False
)

print(
    "\nMerchant vs PPA revenue comparison:"
)

print(
    comparison.to_string(
        index=False
    )
)

# ============================================================
# 16. Export key Phase 1 results to Excel
# ============================================================

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------
# 16.1 Output path
# ------------------------------------------------------------

EXCEL_OUTPUT_PATH = (
    OUTPUT_DIRECTORY
    / "Phase1_Operational_Results.xlsx"
)


# ------------------------------------------------------------
# 16.2 Make sure contract labels exist
# ------------------------------------------------------------

merchant_summary = merchant_summary.copy()

if "contract" not in merchant_summary.columns:
    merchant_summary["contract"] = "Merchant"

if "alpha" not in merchant_summary.columns:
    merchant_summary["alpha"] = 0.0


if "ppa_summary" in globals():

    ppa_summary = ppa_summary.copy()

    if "contract" not in ppa_summary.columns:
        ppa_summary["contract"] = "PPA"

    if "alpha" not in ppa_summary.columns:
        ppa_summary["alpha"] = 1.0


# ------------------------------------------------------------
# 16.3 Function to calculate derived indicators
# ------------------------------------------------------------

def add_analysis_indicators(df):

    df = df.copy()

    hours_in_period = len(model_data)

    # --------------------------------------------------------
    # Grid utilisation rate
    #
    # Annual export /
    # theoretical maximum annual grid export
    # --------------------------------------------------------

    df["Grid_utilisation_rate"] = np.where(
        df["Grid_connection_MW"] > 0,

        df["Grid_export_MWh"]
        /
        (
            df["Grid_connection_MW"]
            * hours_in_period
        ),

        np.nan
    )


    # --------------------------------------------------------
    # Curtailment percentages
    # --------------------------------------------------------

    df["Physical_curtailment_pct"] = (
        df["Physical_curtailment_rate"]
        * 100
    )

    df["Economic_curtailment_pct"] = (
        df["Economic_curtailment_rate"]
        * 100
    )

    df["Total_curtailment_pct"] = (
        df["Total_curtailment_rate"]
        * 100
    )


    # --------------------------------------------------------
    # Revenue per unit of renewable generation
    #
    # BESS-alone has RES_generation = 0,
    # therefore result is NaN.
    # --------------------------------------------------------

    df["Revenue_per_MWh_RES_GBP"] = np.where(
        df["RES_generation_MWh"] > 0,

        df["Net_revenue_GBP"]
        / df["RES_generation_MWh"],

        np.nan
    )


    # --------------------------------------------------------
    # Revenue per MWh exported
    # --------------------------------------------------------

    df["Revenue_per_MWh_export_GBP"] = np.where(
        df["Grid_export_MWh"] > 0,

        df["Net_revenue_GBP"]
        / df["Grid_export_MWh"],

        np.nan
    )


    # --------------------------------------------------------
    # Battery throughput
    #
    # Useful for comparing how heavily the BESS is used.
    # --------------------------------------------------------

    df["Battery_throughput_MWh"] = (
        df["Battery_charge_MWh"]
        + df["Battery_discharge_MWh"]
    )


    # --------------------------------------------------------
    # Approximate equivalent discharged cycles
    #
    # Discharged energy / nominal battery energy capacity
    #
    # Only calculated for BESS scenarios.
    # --------------------------------------------------------

    df["Equivalent_discharge_cycles"] = np.where(
        df["BESS_energy_MWh"] > 0,

        df["Battery_discharge_MWh"]
        / df["BESS_energy_MWh"],

        np.nan
    )


    return df


# ------------------------------------------------------------
# 16.4 Add derived indicators
# ------------------------------------------------------------

merchant_analysis = add_analysis_indicators(
    merchant_summary
)

if "ppa_summary" in globals():

    ppa_analysis = add_analysis_indicators(
        ppa_summary
    )

else:

    ppa_analysis = None


# ------------------------------------------------------------
# 16.5 Combine Merchant and PPA
# ------------------------------------------------------------

if ppa_analysis is not None:

    all_scenarios = pd.concat(
        [
            merchant_analysis,
            ppa_analysis,
        ],
        ignore_index=True
    )

else:

    all_scenarios = merchant_analysis.copy()


all_scenarios = all_scenarios.sort_values(
    by=[
        "configuration",
        "grid_level",
        "contract",
    ]
).reset_index(drop=True)


# ------------------------------------------------------------
# 16.6 Contract comparison
#
# Merchant vs PPA at the same configuration
# and grid connection level
# ------------------------------------------------------------

if ppa_analysis is not None:

    contract_comparison = (
        all_scenarios
        .pivot_table(
            index=[
                "configuration",
                "grid_level",
                "Grid_connection_MW",
            ],
            columns="contract",
            values="Net_revenue_GBP",
            aggfunc="first"
        )
        .reset_index()
    )

    if (
        "Merchant" in contract_comparison.columns
        and "PPA" in contract_comparison.columns
    ):

        contract_comparison[
            "PPA_minus_Merchant_GBP"
        ] = (
            contract_comparison["PPA"]
            - contract_comparison["Merchant"]
        )

        contract_comparison[
            "PPA_vs_Merchant_pct"
        ] = np.where(
            contract_comparison["Merchant"] != 0,

            (
                contract_comparison["PPA"]
                - contract_comparison["Merchant"]
            )
            / contract_comparison["Merchant"]
            * 100,

            np.nan
        )

else:

    contract_comparison = pd.DataFrame()


# ------------------------------------------------------------
# 16.7 Revenue analysis sheet
# ------------------------------------------------------------

revenue_analysis = all_scenarios[
    [
        "contract",
        "configuration",
        "grid_level",
        "Grid_connection_MW",
        "RES_capacity_MW",
        "Grid_export_MWh",
        "Net_revenue_GBP",
        "Revenue_per_MWh_RES_GBP",
        "Revenue_per_MWh_export_GBP",
    ]
].copy()


# ------------------------------------------------------------
# 16.8 Curtailment analysis sheet
# ------------------------------------------------------------

curtailment_analysis = all_scenarios[
    [
        "contract",
        "configuration",
        "grid_level",
        "Grid_connection_MW",
        "RES_capacity_MW",
        "RES_generation_MWh",
        "Grid_export_MWh",
        "Physical_curtailment_MWh",
        "Economic_curtailment_MWh",
        "Total_curtailment_MWh",
        "Physical_curtailment_pct",
        "Economic_curtailment_pct",
        "Total_curtailment_pct",
        "Grid_utilisation_rate",
    ]
].copy()


# ------------------------------------------------------------
# 16.9 Battery analysis sheet
# ------------------------------------------------------------

battery_analysis = all_scenarios[
    all_scenarios["BESS_power_MW"] > 0
][
    [
        "contract",
        "configuration",
        "grid_level",
        "Grid_connection_MW",
        "BESS_power_MW",
        "BESS_energy_MWh",
        "Battery_charge_MWh",
        "Battery_discharge_MWh",
        "Battery_throughput_MWh",
        "Equivalent_discharge_cycles",
        "Net_revenue_GBP",
    ]
].copy()


# ------------------------------------------------------------
# 16.10 Grid sensitivity sheet
#
# Shows how each outcome changes as grid capacity
# increases from 60% to 100%.
# ------------------------------------------------------------

grid_sensitivity = all_scenarios[
    [
        "contract",
        "configuration",
        "grid_level",
        "Grid_connection_MW",
        "Grid_export_MWh",
        "Physical_curtailment_MWh",
        "Economic_curtailment_MWh",
        "Total_curtailment_MWh",
        "Physical_curtailment_pct",
        "Economic_curtailment_pct",
        "Total_curtailment_pct",
        "Grid_utilisation_rate",
        "Battery_discharge_MWh",
        "Net_revenue_GBP",
    ]
].copy()


grid_sensitivity = grid_sensitivity.sort_values(
    by=[
        "contract",
        "configuration",
        "grid_level",
    ]
)


# ------------------------------------------------------------
# 16.11 Model parameters sheet
# ------------------------------------------------------------

parameter_data = pd.DataFrame(
    {
        "Parameter": [
            "Simulation start",
            "Simulation end",
            "Time resolution",
            "PV capacity",
            "Wind capacity",
            "Battery power",
            "Battery energy",
            "Battery duration",
            "Initial SOC",
            "Minimum SOC",
            "Maximum SOC",
            "Charging efficiency",
            "Discharging efficiency",
            "Round-trip efficiency",
            "PPA price",
            "Degradation cost",
            "Merchant alpha",
            "PPA alpha",
            "Grid connection levels",
            "Grid import rule",
            "SOC balancing",
            "Curtailment reporting",
        ],

        "Value": [
            START_DATE,
            END_DATE,
            f"{DELTA_T} hour",
            PV_CAPACITY_MW,
            WIND_CAPACITY_MW,
            BATTERY_POWER_MW,
            BATTERY_ENERGY_MWH,
            BATTERY_ENERGY_MWH / BATTERY_POWER_MW,
            SOC_INITIAL,
            SOC_MIN,
            SOC_MAX,
            ETA_CHARGE,
            ETA_DISCHARGE,
            ETA_CHARGE * ETA_DISCHARGE,
            PPA_PRICE_GBP_PER_MWH,
            DEGRADATION_COST_GBP_PER_MWH,
            0.0,
            1.0,
            "60%, 70%, 80%, 90%, 100%",
            (
                "Grid imports allowed only "
                "for standalone BESS"
            ),
            "Full-horizon terminal SOC balance",
            "Physical + economic + total",
        ],

        "Unit / Description": [
            "UTC",
            "UTC, exclusive upper bound",
            "hours",
            "MW",
            "MW",
            "MW",
            "MWh",
            "hours",
            "fraction",
            "fraction",
            "fraction",
            "fraction",
            "fraction",
            "fraction",
            "GBP/MWh",
            "GBP/MWh discharged",
            "Merchant",
            "100% fixed-price PPA",
            "% of installed RES capacity",
            "P_in = 0 for all other configurations",
            "Final SOC returns to initial SOC",
            "Post-processed decomposition; dispatch unchanged",
        ]
    }
)


# ------------------------------------------------------------
# 16.12 Columns retained in main summary sheets
# ------------------------------------------------------------

summary_columns = [
    "contract",
    "configuration",
    "grid_level",
    "PV_capacity_MW",
    "Wind_capacity_MW",
    "RES_capacity_MW",
    "BESS_power_MW",
    "BESS_energy_MWh",
    "Grid_connection_MW",
    "RES_generation_MWh",
    "Grid_export_MWh",
    "Grid_import_MWh",
    "Physical_curtailment_MWh",
    "Economic_curtailment_MWh",
    "Total_curtailment_MWh",
    "Physical_curtailment_pct",
    "Economic_curtailment_pct",
    "Total_curtailment_pct",
    "Grid_utilisation_rate",
    "Battery_charge_MWh",
    "Battery_discharge_MWh",
    "Battery_throughput_MWh",
    "Equivalent_discharge_cycles",
    "Net_revenue_GBP",
    "Revenue_per_MWh_RES_GBP",
    "Revenue_per_MWh_export_GBP",
]


# ------------------------------------------------------------
# 16.13 Write everything to Excel
# ------------------------------------------------------------

with pd.ExcelWriter(
    EXCEL_OUTPUT_PATH,
    engine="openpyxl"
) as writer:

    # Merchant
    merchant_analysis[
        summary_columns
    ].to_excel(
        writer,
        sheet_name="Merchant Summary",
        index=False
    )

    # PPA
    if ppa_analysis is not None:

        ppa_analysis[
            summary_columns
        ].to_excel(
            writer,
            sheet_name="PPA Summary",
            index=False
        )

    # All scenarios
    all_scenarios[
        summary_columns
    ].to_excel(
        writer,
        sheet_name="All Scenarios",
        index=False
    )

    # Contract comparison
    if not contract_comparison.empty:

        contract_comparison.to_excel(
            writer,
            sheet_name="Contract Comparison",
            index=False
        )

    # Revenue
    revenue_analysis.to_excel(
        writer,
        sheet_name="Revenue Analysis",
        index=False
    )

    # Curtailment
    curtailment_analysis.to_excel(
        writer,
        sheet_name="Curtailment Analysis",
        index=False
    )

    # Battery
    battery_analysis.to_excel(
        writer,
        sheet_name="Battery Analysis",
        index=False
    )

    # Grid sensitivity
    grid_sensitivity.to_excel(
        writer,
        sheet_name="Grid Sensitivity",
        index=False
    )

    # Parameters
    parameter_data.to_excel(
        writer,
        sheet_name="Model Parameters",
        index=False
    )


# ============================================================
# 17. Excel formatting
# ============================================================

workbook = load_workbook(
    EXCEL_OUTPUT_PATH
)

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

for worksheet in workbook.worksheets:

    # Freeze header row
    worksheet.freeze_panes = "A2"

    # Apply autofilter
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    # Header formatting
    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    # Automatically adjust column width
    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is not None:

                cell_length = len(
                    str(cell.value)
                )

                if cell_length > max_length:
                    max_length = cell_length

        adjusted_width = min(
            max(max_length + 2, 12),
            28
        )

        worksheet.column_dimensions[
            column_letter
        ].width = adjusted_width


    # Number formatting
    for row in worksheet.iter_rows(
        min_row=2
    ):

        for cell in row:

            header = worksheet.cell(
                row=1,
                column=cell.column
            ).value

            if header is None:
                continue

            # Currency
            if (
                "GBP" in str(header)
                or "revenue" in str(header).lower()
            ):

                cell.number_format = (
                    '£#,##0.00'
                )

            # Percentages
            elif header in [
                "grid_level",
                "Grid_utilisation_rate",
            ]:

                cell.number_format = (
                    "0.0%"
                )

            elif header in [
                "Physical_curtailment_pct",
                "Economic_curtailment_pct",
                "Total_curtailment_pct",
                "PPA_vs_Merchant_pct",
            ]:

                cell.number_format = (
                    '0.00"%"'
                )

            # Energy/capacity
            elif (
                "_MW" in str(header)
                or "_MWh" in str(header)
            ):

                cell.number_format = (
                    "#,##0.00"
                )

            elif header == (
                "Equivalent_discharge_cycles"
            ):

                cell.number_format = (
                    "#,##0.00"
                )


# Save formatted workbook
workbook.save(
    EXCEL_OUTPUT_PATH
)


# ============================================================
# 18. Final checks / console output
# ============================================================

print(
    "\n================================================"
)

print(
    "EXCEL EXPORT COMPLETED"
)

print(
    "================================================"
)

print(
    f"\nExcel file saved to:\n"
    f"{EXCEL_OUTPUT_PATH}"
)

print(
    "\nSheets created:"
)

for sheet_name in workbook.sheetnames:

    print(
        f" - {sheet_name}"
    )

print(
    f"\nTotal scenarios exported: "
    f"{len(all_scenarios)}"
)
